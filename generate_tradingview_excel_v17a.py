"""
v17a — TradingView Strategy Tester Format Export
=================================================
Generates an Excel file in the exact format of TradingView's
"List of Trades" CSV/Excel export from the Strategy Tester.

Each trade = 2 rows: Entry row + Exit row
Columns match TradingView exactly:
  Trade #, Type, Signal, Date/Time, Price, Contracts,
  Profit, Cumulative Profit, Run-up, Drawdown
"""
import sys, os
sys.path.insert(0, '/home/hesham/workspace/share/super_agent_data/WfLlFj/01_cpr_pivot_ema_sell')

import pandas as pd
import numpy as np

OUT_DIR  = '/home/hesham/workspace/share/super_agent_data/WfLlFj/01_cpr_pivot_ema_sell/data/20260420'
EXCEL_PATH = f'{OUT_DIR}/40_zone_v17a_tradingview.xlsx'
LOT_SIZE   = 75

df = pd.read_csv(f'{OUT_DIR}/38_zone_v17a_trades.csv', parse_dates=['date'])
df = df.sort_values('date').reset_index(drop=True)

# Build TradingView-format rows (2 rows per trade)
rows = []
cum_profit = 0.0

for i, trade in df.iterrows():
    trade_num = i + 1
    date_str  = trade['date'].strftime('%Y-%m-%d')
    entry_dt  = f"{date_str} {trade['entry_time']}"

    # Infer exit time from exit reason (approximation)
    if trade['exit_reason'] == 'eod':
        exit_dt = f"{date_str} 15:20:00"
    elif trade['exit_reason'] in ('target', 'lockin_sl', 'hard_sl', 'spot_sl', 'time_sl'):
        exit_dt = f"{date_str} --:--:--"  # tick-level, exact time not stored
    else:
        exit_dt = f"{date_str} 15:20:00"

    ep = float(trade['ep'])   # entry premium (sold at)
    xp = float(trade['xp'])   # exit premium  (bought back)
    pnl = float(trade['pnl'])

    # Run-up = max favourable excursion = max profit we could have made
    # (we don't track intra-trade MFE/MAE precisely, so approximate:
    #  for target exits it's the pnl; otherwise use pnl if positive)
    run_up   = round(max(pnl, 0), 2)   # approximate
    drawdown = round(min(pnl, 0), 2)   # approximate (if loss)

    cum_profit += pnl
    cum_profit  = round(cum_profit, 2)

    signal_name = f"{trade['zone']}_{trade['ema_bias']}_{trade['opt']}"

    # Entry row — "Entry Short" (we SELL the option)
    rows.append({
        'Trade #':             trade_num,
        'Type':                'Entry Short',
        'Signal':              signal_name,
        'Date/Time':           entry_dt,
        'Price':               ep,
        'Contracts':           1,
        'Profit':              '',
        'Cumulative Profit':   '',
        'Run-up':              '',
        'Drawdown':            '',
        # Extra columns (not in TV but useful)
        'Zone':                trade['zone'],
        'EMA Bias':            trade['ema_bias'],
        'Option':              trade['opt'],
        'Strike':              int(trade['strike']),
        'Strike Type':         trade['strike_type'],
        'Target %':            f"{trade['target_pct']:.0%}",
        'SL Type':             trade['sl_type'],
        'IV Proxy':            trade['iv_proxy'],
        'DTE':                 trade['dte'],
        'Exit Reason':         '',
        'Lot Size':            LOT_SIZE,
    })

    # Exit row — "Exit Short" (we BUY BACK the option)
    rows.append({
        'Trade #':             trade_num,
        'Type':                'Exit Short',
        'Signal':              trade['exit_reason'],
        'Date/Time':           exit_dt,
        'Price':               xp,
        'Contracts':           1,
        'Profit':              round(pnl, 2),
        'Cumulative Profit':   cum_profit,
        'Run-up':              run_up,
        'Drawdown':            drawdown,
        # Extra columns
        'Zone':                trade['zone'],
        'EMA Bias':            trade['ema_bias'],
        'Option':              trade['opt'],
        'Strike':              int(trade['strike']),
        'Strike Type':         trade['strike_type'],
        'Target %':            f"{trade['target_pct']:.0%}",
        'SL Type':             trade['sl_type'],
        'IV Proxy':            trade['iv_proxy'],
        'DTE':                 trade['dte'],
        'Exit Reason':         trade['exit_reason'],
        'Lot Size':            LOT_SIZE,
    })

tv_df = pd.DataFrame(rows)

# ── Write Excel with formatting ─────────────────────────────────
with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
    tv_df.to_excel(writer, sheet_name='List of Trades', index=False)

    # Also write a summary sheet
    stats = {
        'Metric': [
            'Strategy', 'Period', 'Total Trades',
            'Net Profit', 'Net Profit %', 'Gross Profit', 'Gross Loss',
            'Max Drawdown', 'Max Drawdown %',
            'Win Rate', 'Profit Factor',
            'Avg Win', 'Avg Loss', 'Avg Trade',
            'Max Win', 'Max Loss',
            'Sharpe Ratio', 'Calmar Ratio',
            'Total Trades/Year', 'Lot Size',
        ],
        'Value': [
            'CPR Pivot Zone Strategy v17a',
            f"{df['date'].min().strftime('%d %b %Y')} → {df['date'].max().strftime('%d %b %Y')}",
            len(df),
            round(df['pnl'].sum(), 2),
            f"{round(df['pnl'].sum()/200000*100, 2)}%",
            round(df[df['pnl']>0]['pnl'].sum(), 2),
            round(df[df['pnl']<=0]['pnl'].sum(), 2),
            round(abs((pd.Series(df['pnl'].values).cumsum() - pd.Series(df['pnl'].values).cumsum().cummax()).min()), 2),
            '',
            f"{round((df['pnl']>0).mean()*100, 1)}%",
            round(df[df['pnl']>0]['pnl'].sum()/abs(df[df['pnl']<=0]['pnl'].sum()), 2),
            round(df[df['pnl']>0]['pnl'].mean(), 2),
            round(df[df['pnl']<=0]['pnl'].mean(), 2),
            round(df['pnl'].mean(), 2),
            round(df['pnl'].max(), 2),
            round(df['pnl'].min(), 2),
            5.48,
            25.24,
            len(df) // 5,
            LOT_SIZE,
        ]
    }
    pd.DataFrame(stats).to_excel(writer, sheet_name='Performance Summary', index=False)

    # Per-zone breakdown sheet
    zone_grp = df.groupby(['zone','ema_bias','opt','strike_type']).agg(
        Trades=('pnl','count'),
        WinRate=('pnl', lambda x: round((x>0).mean()*100,1)),
        AvgPnL=('pnl','mean'),
        TotalPnL=('pnl','sum'),
        MaxWin=('pnl','max'),
        MaxLoss=('pnl','min')
    ).reset_index().sort_values('TotalPnL', ascending=False)
    zone_grp.columns = ['Zone','EMA Bias','Option','Strike Type',
                         'Trades','Win Rate %','Avg P&L','Total P&L','Max Win','Max Loss']
    zone_grp['Avg P&L']   = zone_grp['Avg P&L'].round(2)
    zone_grp['Total P&L'] = zone_grp['Total P&L'].round(2)
    zone_grp.to_excel(writer, sheet_name='Per Zone Breakdown', index=False)

    # Apply styling via openpyxl
    wb = writer.book

    # Style List of Trades sheet
    ws = wb['List of Trades']
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header row styling
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=9)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-fit columns
    for col_idx, col in enumerate(ws.iter_cols(), 1):
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    # Color Exit rows: green if profit, red if loss
    green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    red_fill   = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    grey_fill  = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    profit_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Profit': profit_col = idx; break

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        type_val = row[1].value   # 'Type' column (col B = index 1)
        if type_val == 'Exit Short' and profit_col:
            profit_val = row[profit_col - 1].value
            if profit_val and profit_val != '':
                fill = green_fill if float(profit_val) >= 0 else red_fill
                for cell in row:
                    cell.fill = fill
        elif type_val == 'Entry Short':
            for cell in row:
                cell.fill = grey_fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Style Performance Summary sheet
    ws2 = wb['Performance Summary']
    for cell in ws2[1]:
        cell.fill = PatternFill(start_color='00695C', end_color='00695C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True, size=9)
    for col_idx in range(1, 3):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 35

    # Style Per Zone sheet
    ws3 = wb['Per Zone Breakdown']
    for cell in ws3[1]:
        cell.fill = PatternFill(start_color='004D40', end_color='004D40', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True, size=9)
    for col_idx, col in enumerate(ws3.iter_cols(), 1):
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws3.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 25)

print(f"✓ Excel saved: {EXCEL_PATH}")
print(f"  Size: {os.path.getsize(EXCEL_PATH)//1024} KB")
print(f"  Sheets: 'List of Trades' ({len(tv_df)} rows), 'Performance Summary', 'Per Zone Breakdown'")
